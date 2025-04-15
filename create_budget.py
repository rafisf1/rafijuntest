import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE

# ----------------------------
# Create a new workbook object
# ----------------------------
wb = Workbook()

# Define styles
header_font = Font(name='Arial', size=11, bold=True)
header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
currency_format = FORMAT_CURRENCY_USD_SIMPLE
percent_format = FORMAT_PERCENTAGE

# Define the worksheet names (order as in the solved example)
sheet_names = [
    "Assumptions",
    "Sales Budget",
    "Cash Collections",
    "Merchandise Purchases",
    "Purchases Disbursements",
    "Selling & Admin Expenses",
    "Cash Budget",
    "Income Statement",
    "Balance Sheet"
]

# Rename default sheet as the first sheet: "Assumptions"
ws_assumptions = wb.active
ws_assumptions.title = sheet_names[0]

# Create the remaining sheets
for name in sheet_names[1:]:
    wb.create_sheet(title=name)

# Function to apply header formatting
def format_header_row(worksheet, row_num):
    for cell in worksheet[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

# Function to apply data formatting
def format_data_row(worksheet, row_num, currency_cols=None, percent_cols=None):
    for col_num, cell in enumerate(worksheet[row_num], 1):
        cell.border = border
        cell.alignment = Alignment(horizontal='right' if isinstance(cell.value, (int, float)) else 'left')
        if currency_cols and col_num in currency_cols:
            cell.number_format = currency_format
        if percent_cols and col_num in percent_cols:
            cell.number_format = percent_format

# Function to format section headers
def format_section_header(worksheet, row_num):
    cell = worksheet.cell(row=row_num, column=1)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')

# -----------------------------
# 1. ASSUMPTIONS SHEET
# -----------------------------
assumptions_data = [
    ["Assumption", "Value", "Description"],
    ["Sales Price per Unit", 50, "Retail price for one unit"],
    ["Current Month Collection %", 0.40, "Percent of current month sales collected in the month of sale"],
    ["Next Month Collection %", 0.60, "Percent of sales collected in the following month"],
    ["COGS Percentage", 0.60, "Percentage of sales that is cost of goods sold"],
    ["Desired Ending Inventory %", 0.20, "Percent of next month COGS to carry as ending inventory"],
    ["Payment This Month %", 0.50, "Percent of purchase paid in current month"],
    ["Payment Next Month %", 0.50, "Percent of purchase paid in the following month"],
    ["Monthly Fixed S&A Expense", 3000, "Fixed selling & admin expense per month"],
    ["Depreciation Expense", 500, "Monthly depreciation expense"],
    ["Minimum Cash Balance", 5000, "Minimum cash to be maintained"],
    ["Beginning Cash", 10000, "Cash at beginning (January)"],
    ["Beginning Inventory", 6000, "Merchandise inventory at start"],
    ["Beginning A/R", 8000, "Accounts receivable balance from December"],
    ["Beginning A/P", 3000, "Accounts payable balance from December"],
    ["Beginning Retained Earnings", 10000, "Retained earnings at start"],
    ["Assumed Equipment Cost", 40000, "Cost of equipment (not adjusted for depreciation yet)"]
]

for row in assumptions_data:
    ws_assumptions.append(row)

# Format Assumptions sheet
format_header_row(ws_assumptions, 1)
for row in range(2, len(assumptions_data) + 1):
    format_data_row(ws_assumptions, row, currency_cols=[2], percent_cols=[2])

# Adjust column widths
for col in range(1, 4):
    ws_assumptions.column_dimensions[get_column_letter(col)].width = 30

# -----------------------------
# 2. SALES BUDGET SHEET
# -----------------------------
ws_sales = wb["Sales Budget"]
ws_sales.append(["Month", "Sales Volume", "Unit Price", "Sales Revenue"])
ws_sales.append(["January 2026", 1000, 50, 1000 * 50])
ws_sales.append(["February 2026", 1200, 50, 1200 * 50])
ws_sales.append(["March 2026", 1500, 50, 1500 * 50])
ws_sales.append(["Q1 Total", "=B2+B3+B4", "", "=D2+D3+D4"])

# Format Sales Budget sheet
format_header_row(ws_sales, 1)
for row in range(2, 6):
    format_data_row(ws_sales, row, currency_cols=[3, 4])

for col in range(1, 5):
    ws_sales.column_dimensions[get_column_letter(col)].width = 20

# -----------------------------
# 3. CASH COLLECTIONS SHEET
# -----------------------------
ws_cc = wb["Cash Collections"]
ws_cc.append(["Month", "From Prior Month", "From Current Sales", "Total Collections"])
ws_cc.append(["January 2026", 8000, 0.40 * 50000, 8000 + 0.40 * 50000])
ws_cc.append(["February 2026", 0.60 * 50000, 0.40 * 60000, 0.60 * 50000 + 0.40 * 60000])
ws_cc.append(["March 2026", 0.60 * 60000, 0.40 * 75000, 0.60 * 60000 + 0.40 * 75000])

# Format Cash Collections sheet
format_header_row(ws_cc, 1)
for row in range(2, 5):
    format_data_row(ws_cc, row, currency_cols=[2, 3, 4])

for col in range(1, 5):
    ws_cc.column_dimensions[get_column_letter(col)].width = 25

# -----------------------------
# 4. MERCHANDISE PURCHASES SHEET
# -----------------------------
ws_mp = wb["Merchandise Purchases"]
ws_mp.append(["Month", "COGS", "Desired Ending Inventory", "Total Needed", "Beginning Inventory", "Purchases"])
ws_mp.append(["January 2026", 30000, 7200, 37200, 6000, 37200 - 6000])
ws_mp.append(["February 2026", 36000, 9000, 45000, 7200, 45000 - 7200])
ws_mp.append(["March 2026", 45000, 9000, 54000, 9000, 54000 - 9000])

# Format Merchandise Purchases sheet
format_header_row(ws_mp, 1)
for row in range(2, 5):
    format_data_row(ws_mp, row, currency_cols=[2, 3, 4, 5, 6])

for col in range(1, 7):
    ws_mp.column_dimensions[get_column_letter(col)].width = 25

# -----------------------------
# 5. PURCHASES DISBURSEMENTS SHEET
# -----------------------------
ws_pd = wb["Purchases Disbursements"]
ws_pd.append(["Month", "From Prior Month", "From Current Purchases", "Total Disbursements"])
ws_pd.append(["January 2026", 3000, 0.5 * 31200, 3000 + 0.5 * 31200])
ws_pd.append(["February 2026", 0.5 * 31200, 0.5 * 37800, 0.5 * 31200 + 0.5 * 37800])
ws_pd.append(["March 2026", 0.5 * 37800, 0.5 * 45000, 0.5 * 37800 + 0.5 * 45000])

# Format Purchases Disbursements sheet
format_header_row(ws_pd, 1)
for row in range(2, 5):
    format_data_row(ws_pd, row, currency_cols=[2, 3, 4])

for col in range(1, 5):
    ws_pd.column_dimensions[get_column_letter(col)].width = 25

# -----------------------------
# 6. SELLING & ADMIN EXPENSES SHEET
# -----------------------------
ws_sa = wb["Selling & Admin Expenses"]
ws_sa.append(["Month", "Fixed Expense", "Depreciation", "Total S&A Expense"])
ws_sa.append(["January 2026", 3000, 500, 3500])
ws_sa.append(["February 2026", 3000, 500, 3500])
ws_sa.append(["March 2026", 3000, 500, 3500])

# Format Selling & Admin Expenses sheet
format_header_row(ws_sa, 1)
for row in range(2, 5):
    format_data_row(ws_sa, row, currency_cols=[2, 3, 4])

for col in range(1, 5):
    ws_sa.column_dimensions[get_column_letter(col)].width = 25

# -----------------------------
# 7. CASH BUDGET SHEET
# -----------------------------
ws_cash = wb["Cash Budget"]
ws_cash.append(["Month", "Beginning Cash", "Collections", "Total Cash Available", "Disbursements", "Borrowing", "Repayment", "Interest", "Ending Cash"])

# January
begin_cash_jan = 10000
collections_jan = 28000
total_available_jan = begin_cash_jan + collections_jan
disbursements_jan = 18600 + 3500
ending_cash_jan = total_available_jan - disbursements_jan
ws_cash.append(["January 2026", begin_cash_jan, collections_jan, total_available_jan, disbursements_jan, 0, 0, 0, ending_cash_jan])

# February
begin_cash_feb = ending_cash_jan
collections_feb = 54000
total_available_feb = begin_cash_feb + collections_feb
disbursements_feb = 34500 + 3500
ending_cash_feb = total_available_feb - disbursements_feb
ws_cash.append(["February 2026", begin_cash_feb, collections_feb, total_available_feb, disbursements_feb, 0, 0, 0, ending_cash_feb])

# March
begin_cash_mar = ending_cash_feb
collections_mar = 66000
total_available_mar = begin_cash_mar + collections_mar
disbursements_mar = 41400 + 3500
ending_cash_mar = total_available_mar - disbursements_mar
ws_cash.append(["March 2026", begin_cash_mar, collections_mar, total_available_mar, disbursements_mar, 0, 0, 0, ending_cash_mar])

# Format Cash Budget sheet
format_header_row(ws_cash, 1)
for row in range(2, 5):
    format_data_row(ws_cash, row, currency_cols=[2, 3, 4, 5, 6, 7, 8, 9])

for col in range(1, 10):
    ws_cash.column_dimensions[get_column_letter(col)].width = 20

# -----------------------------
# 8. INCOME STATEMENT SHEET
# -----------------------------
ws_is = wb["Income Statement"]
ws_is.append(["Item", "Q1 Total"])

# Calculate totals
total_sales = 50000 + 60000 + 75000
total_COGS = 30000 + 36000 + 45000
gross_profit = total_sales - total_COGS
total_SA = 3500 * 3
operating_income = gross_profit - total_SA

ws_is.append(["Sales Revenue", total_sales])
ws_is.append(["COGS", total_COGS])
ws_is.append(["Gross Profit", gross_profit])
ws_is.append(["Selling & Admin Expense", total_SA])
ws_is.append(["Operating Income", operating_income])
ws_is.append(["Interest Expense", 0])
ws_is.append(["Net Income", operating_income])

# Format Income Statement sheet
format_header_row(ws_is, 1)
for row in range(2, 9):
    format_data_row(ws_is, row, currency_cols=[2])

for col in range(1, 3):
    ws_is.column_dimensions[get_column_letter(col)].width = 30

# -----------------------------
# 9. BALANCE SHEET SHEET
# -----------------------------
ws_bs = wb["Balance Sheet"]
ws_bs.append(["Item", "Ending Balance"])

# Assets
ws_bs.append(["--- Assets ---", ""])
ws_bs.append(["Cash", ending_cash_mar])
ws_bs.append(["Accounts Receivable", 0.60 * 75000])
ws_bs.append(["Inventory", 9000])
ws_bs.append(["Equipment (Net)", 40000 - 1500])
total_assets = ending_cash_mar + (0.60 * 75000) + 9000 + (40000 - 1500)
ws_bs.append(["Total Assets", total_assets])

# Liabilities
ws_bs.append(["", ""])
ws_bs.append(["--- Liabilities ---", ""])
ws_bs.append(["Accounts Payable", 0.5 * 45000])
ws_bs.append(["Total Liabilities", 0.5 * 45000])

# Equity
ws_bs.append(["", ""])
ws_bs.append(["--- Equity ---", ""])
retained_earnings = 10000 + operating_income
total_equity = total_assets - (0.5 * 45000)
common_stock = total_equity - retained_earnings

ws_bs.append(["Common Stock", common_stock])
ws_bs.append(["Retained Earnings", retained_earnings])
ws_bs.append(["Total Equity", common_stock + retained_earnings])
ws_bs.append(["Total Liabilities & Equity", (0.5 * 45000) + common_stock + retained_earnings])

# Format Balance Sheet sheet
format_header_row(ws_bs, 1)
format_section_header(ws_bs, 3)  # Assets header
format_section_header(ws_bs, 9)  # Liabilities header
format_section_header(ws_bs, 12)  # Equity header

for row in range(2, 16):
    if row not in [3, 9, 12]:  # Skip section headers
        format_data_row(ws_bs, row, currency_cols=[2])

for col in range(1, 3):
    ws_bs.column_dimensions[get_column_letter(col)].width = 30

# Save the workbook
file_path = "The_Iggy_Store_Master_Budget_v3.xlsx"
wb.save(file_path)
print("Workbook created and saved as:", file_path) 